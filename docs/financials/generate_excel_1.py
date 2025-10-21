import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.chart import PieChart, BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList

def create_excel_workbook():
    wb = openpyxl.Workbook()

    # --- Styles ---
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    native_coin_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    bold_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    currency_format = '"$"#,##0.00'
    percentage_format = '0.00%'
    integer_format = '#,##0'

    # --- Inputs Sheet ---
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"

    # Title
    ws_inputs['A1'] = "AEQUITAS PROTOCOL - FINANCIAL INPUTS"
    ws_inputs['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws_inputs.merge_cells('A1:E1')
    ws_inputs['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Asset Classification Banner
    ws_inputs['A2'] = "REPAR: NATIVE COIN OF SOVEREIGN AI JUSTICE BLOCKCHAIN"
    ws_inputs['A2'].font = Font(bold=True, size=11, color='006100')
    ws_inputs['A2'].fill = native_coin_fill
    ws_inputs.merge_cells('A2:E2')
    ws_inputs['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # Key Financial Metrics
    ws_inputs['A4'] = "Key Financial Metrics"
    ws_inputs['A4'].font = Font(bold=True, size=14, color="1F4E79")
    ws_inputs.merge_cells('A4:E4')

    headers_metrics = ["Metric", "Value", "Unit", "Source/Notes"]
    ws_inputs.append(headers_metrics)
    for col_idx, header in enumerate(headers_metrics):
        cell = ws_inputs.cell(row=5, column=col_idx + 1)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    data_metrics = [
        ["Development Cost", 28000000, "USD", "Tri-Agent Validated Consensus"],
        ["Pre-Launch Valuation", 7000000000, "USD", "Sovereign Blockchain + AI Protocol"],
        ["  - Blockchain Infrastructure", 2500000000, "USD", "Base layer sovereign chain value"],
        ["  - AI Protocol Value", 3000000000, "USD", "Cerberus + enforcement mechanisms"],
        ["  - Native Coin Economics", 1000000000, "USD", "Deflationary model + utility"],
        ["  - Network Effects", 500000000, "USD", "150K+ descendant community"],
        ["After-Launch (Y1) Valuation", 250000000000, "USD", "Native Coin Model (Updated)"],
        ["Operational War Chest", 22000000, "USD", "18-month enforcement runway"],
        ["Total Addressable Market (TAM)", 131000000000000, "USD", "$131T Proven Liability"]
    ]
    for row_idx, row_data in enumerate(data_metrics, start=6):
        ws_inputs.append(row_data)
        for col_idx, cell_value in enumerate(row_data):
            cell = ws_inputs.cell(row=row_idx, column=col_idx + 1)
            cell.border = thin_border
            if col_idx == 1:
                cell.number_format = currency_format
            if row_data[0].startswith("  -"):
                cell.font = Font(italic=True, size=10)

    # Investment Details
    ws_inputs['A15'] = "Investment Details"
    ws_inputs['A15'].font = Font(bold=True, size=14, color="1F4E79")
    ws_inputs.merge_cells('A15:E15')

    headers_investment = ["Metric", "Value", "Unit", "Source/Notes"]
    ws_inputs.append(headers_investment)
    for col_idx, header in enumerate(headers_investment):
        cell = ws_inputs.cell(row=16, column=col_idx + 1)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    data_investment = [
        ["Seed Raise", 22000000, "USD", ""],
        ["Pre-Money Valuation", 7000000000, "USD", ""],
        ["Equity Percentage", "", "%", "Calculated: Seed Raise / Pre-Money"],
        ["Implied Value Per $1", "", "USD", "Pre-Money / Seed Raise"]
    ]
    for row_idx, row_data in enumerate(data_investment, start=17):
        ws_inputs.append(row_data)
        for col_idx, cell_value in enumerate(row_data):
            cell = ws_inputs.cell(row=row_idx, column=col_idx + 1)
            cell.border = thin_border
            if col_idx == 1:
                if row_data[0] == "Equity Percentage":
                    cell.number_format = percentage_format
                    cell.value = '=IFERROR(B17/B18,0)'
                elif row_data[0] == "Implied Value Per $1":
                    cell.number_format = currency_format
                    cell.value = '=IFERROR(B18/B17,0)'
                else:
                    cell.number_format = currency_format

    # Use of Funds
    ws_inputs['A22'] = "Use of Funds - 18 Month Runway"
    ws_inputs['A22'].font = Font(bold=True, size=14, color="1F4E79")
    ws_inputs.merge_cells('A22:E22')

    headers_funds = ["Category", "Amount", "% of Raise", "Justification"]
    ws_inputs.append(headers_funds)
    for col_idx, header in enumerate(headers_funds):
        cell = ws_inputs.cell(row=23, column=col_idx + 1)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    data_funds = [
        ["Legal & Enforcement", 7500000, "", "Arbitral Swarm against defendants"],
        ["Security Operations", 5000000, "", "State-level protection"],
        ["Elite Core Team", 3000000, "", "AI engineers, legal strategists"],
        ["AI Infrastructure", 2000000, "", "Cerberus compute, training"],
        ["Contingency Reserve", 4500000, "", "Defense fund for counter-attacks"]
    ]
    for row_idx, row_data in enumerate(data_funds, start=24):
        ws_inputs.append(row_data)
        for col_idx, cell_value in enumerate(row_data):
            cell = ws_inputs.cell(row=row_idx, column=col_idx + 1)
            cell.border = thin_border
            if col_idx == 1:
                cell.number_format = currency_format
            elif col_idx == 2:
                cell.number_format = percentage_format
                cell.value = f"=IFERROR(B{row_idx}/$B$17,0)"

    ws_inputs['A29'] = "TOTAL"
    ws_inputs['B29'].value = "=SUM(B24:B28)"
    ws_inputs['B29'].number_format = currency_format
    ws_inputs['C29'].value = "=SUM(C24:C28)"
    ws_inputs['C29'].number_format = percentage_format
    for col in ['A', 'B', 'C']:
        ws_inputs[col + '29'].font = Font(bold=True)
        ws_inputs[col + '29'].border = thin_border

    # Native Coin Revenue Streams
    ws_inputs['A31'] = "Native Coin Revenue Streams (REPAR Advantage)"
    ws_inputs['A31'].font = Font(bold=True, size=12, color='006100')
    ws_inputs['A31'].fill = native_coin_fill
    ws_inputs.merge_cells('A31:E31')

    headers_revenue = ["Revenue Stream", "Year 1 Potential", "Year 3 Potential", "Description"]
    ws_inputs.append(headers_revenue)
    for col_idx, header in enumerate(headers_revenue):
        cell = ws_inputs.cell(row=32, column=col_idx + 1)
        cell.font = header_font
        cell.fill = native_coin_fill
        cell.border = thin_border

    data_revenue = [
        ["Transaction Fees", 500000000, 5000000000, "All network gas in REPAR"],
        ["Validator Economics", 200000000, 2000000000, "Staking + node licensing"],
        ["Cross-Chain Bridges", 100000000, 1000000000, "Gateway monopoly fees"],
        ["Justice Enforcement", 10000000000, 100000000000, "Settlement recovery"]
    ]
    for row_idx, row_data in enumerate(data_revenue, start=33):
        ws_inputs.append(row_data)
        for col_idx, cell_value in enumerate(row_data):
            cell = ws_inputs.cell(row=row_idx, column=col_idx + 1)
            cell.border = thin_border
            if col_idx in [1, 2]:
                cell.number_format = currency_format

    # Return Projections
    ws_inputs['A38'] = "Return Projections - Native Coin Model"
    ws_inputs['A38'].font = Font(bold=True, size=14, color="1F4E79")
    ws_inputs.merge_cells('A38:E38')

    headers_returns = ["Scenario", "Year 1 MC", "Year 3 MC", "Return Multiple"]
    ws_inputs.append(headers_returns)
    for col_idx, header in enumerate(headers_returns):
        cell = ws_inputs.cell(row=39, column=col_idx + 1)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    data_returns = [
        ["Conservative", 150000000000, 750000000000, 21],
        ["Expected", 250000000000, 1500000000000, 43],
        ["Aggressive", 400000000000, 3500000000000, 50],
        ["Paradigm Shift", 600000000000, 7000000000000, 100]
    ]
    for row_idx, row_data in enumerate(data_returns, start=40):
        ws_inputs.append(row_data)
        for col_idx, cell_value in enumerate(row_data):
            cell = ws_inputs.cell(row=row_idx, column=col_idx + 1)
            cell.border = thin_border
            if col_idx in [1, 2]:
                cell.number_format = currency_format
            elif col_idx == 3:
                cell.number_format = '0"x"'

    # Comparable Analysis
    ws_inputs['A45'] = "Comparable Analysis - Native Coin vs Token"
    ws_inputs['A45'].font = Font(bold=True, size=14, color="1F4E79")
    ws_inputs.merge_cells('A45:E45')

    headers_comparables = ["Protocol", "Type", "Function", "Market Cap", "TAM"]
    ws_inputs.append(headers_comparables)
    for col_idx, header in enumerate(headers_comparables):
        cell = ws_inputs.cell(row=46, column=col_idx + 1)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    data_comparables = [
        ["Ethereum", "Native Coin", "Smart Contracts", 220000000000, 10000000000000],
        ["Chainlink", "Token", "Oracle", 8000000000, 1000000000000],
        ["Cosmos", "Native Coin", "Interoperability", 2500000000, 5000000000000],
        ["Aequitas", "Native Coin", "AI Justice", 7000000000, 131000000000000]
    ]
    for row_idx, row_data in enumerate(data_comparables, start=47):
        ws_inputs.append(row_data)
        for col_idx, cell_value in enumerate(row_data):
            cell = ws_inputs.cell(row=row_idx, column=col_idx + 1)
            cell.border = thin_border
            if col_idx in [3, 4]:
                cell.number_format = currency_format
            if row_data[1] == "Native Coin":
                cell.fill = native_coin_fill

    # Sensitivity Analysis Inputs
    ws_inputs['A52'] = "Sensitivity Analysis - Collection Rate vs Returns"
    ws_inputs['A52'].font = Font(bold=True, size=14, color="1F4E79")
    ws_inputs.merge_cells('A52:E52')

    headers_sensitivity = ["Collection Rate (%)", "Recovered Amount", "Return Multiple (Est.)"]
    ws_inputs.append(headers_sensitivity)
    for col_idx, header in enumerate(headers_sensitivity):
        cell = ws_inputs.cell(row=53, column=col_idx + 1)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    sensitivity_rates = [0.01, 0.05, 0.1, 0.5, 1.0, 5.0]  # % rates
    for row_idx, rate in enumerate(sensitivity_rates, start=54):
        ws_inputs[f'A{row_idx}'] = rate / 100
        ws_inputs[f'A{row_idx}'].number_format = percentage_format
        ws_inputs[f'B{row_idx}'].value = f"=A{row_idx}*B14"  # TAM is at B14
        ws_inputs[f'B{row_idx}'].number_format = currency_format
        ws_inputs[f'C{row_idx}'].value = f"=IFERROR((A{row_idx}/0.0005)*29,0)"  # Scale from 0.05% base
        ws_inputs[f'C{row_idx}'].number_format = '0"x"'
        for col_letter in ['A', 'B', 'C']:
            ws_inputs[f'{col_letter}{row_idx}'].border = thin_border

    # Auto-fit columns
    for col in ws_inputs.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_inputs.column_dimensions[column].width = adjusted_width

    # Freeze panes
    ws_inputs.freeze_panes = 'A6' # Freeze rows above the first data set

    # --- Calculations Sheet ---
    ws_calcs = wb.create_sheet("Calculations")
    ws_calcs['A1'] = "AEQUITAS PROTOCOL - CALCULATIONS"
    ws_calcs['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws_calcs.merge_cells('A1:E1')
    ws_calcs['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Define Calculation Sections
    calc_sections = [
        {"title": "Investment Ratios", "start_row": 3, "headers": ["Metric", "Value", "Formula", "Notes"]},
        {"title": "Use of Funds Summary", "start_row": 10, "headers": ["Category", "Amount", "% of Total Raise", "Notes"]},
        {"title": "Revenue Projections Summary", "start_row": 20, "headers": ["Metric", "Year 1 Total", "Year 3 Total", "Growth (Y1 to Y3)"]},
        {"title": "Return Analysis", "start_row": 30, "headers": ["Scenario", "Equity Value (Y1)", "Equity Value (Y3)", "Return Multiple (Calculated)"]}
    ]

    current_row = 3
    for section in calc_sections:
        # Section Title
        ws_calcs[f'A{current_row}'] = section["title"]
        ws_calcs[f'A{current_row}'].font = Font(bold=True, size=14, color="1F4E79")
        ws_calcs.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1

        # Headers
        ws_calcs.append(section["headers"])
        for col_idx, header in enumerate(section["headers"]):
            cell = ws_calcs.cell(row=current_row, column=col_idx + 1)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        current_row += 1

        # Data and Formulas for each section
        if section["title"] == "Investment Ratios":
            ws_calcs[f'A{current_row}'] = "Seed Raise"
            ws_calcs[f'B{current_row}'].value = '=Inputs!B17'
            ws_calcs[f'B{current_row}'].number_format = currency_format
            ws_calcs[f'C{current_row}'] = "Inputs!B17"
            ws_calcs[f'D{current_row}'] = "Direct from Inputs"
            for col_letter in ['A', 'B', 'C', 'D']:
                 ws_calcs[f'{col_letter}{current_row}'].border = thin_border
            current_row += 1

            ws_calcs[f'A{current_row}'] = "Pre-Money Valuation"
            ws_calcs[f'B{current_row}'].value = '=Inputs!B18'
            ws_calcs[f'B{current_row}'].number_format = currency_format
            ws_calcs[f'C{current_row}'] = "Inputs!B18"
            ws_calcs[f'D{current_row}'] = "Direct from Inputs"
            for col_letter in ['A', 'B', 'C', 'D']:
                 ws_calcs[f'{col_letter}{current_row}'].border = thin_border
            current_row += 1

            ws_calcs[f'A{current_row}'] = "Equity Percentage (Calculated)"
            ws_calcs[f'B{current_row}'].value = f'=IFERROR(B{current_row-2}/B{current_row-1},0)'
            ws_calcs[f'B{current_row}'].number_format = percentage_format
            ws_calcs[f'C{current_row}'] = "Seed Raise / Pre-Money Valuation"
            ws_calcs[f'D{current_row}'] = "Calculated from Inputs"
            for col_letter in ['A', 'B', 'C', 'D']:
                 ws_calcs[f'{col_letter}{current_row}'].border = thin_border
            current_row += 1

            ws_calcs[f'A{current_row}'] = "Implied Value Per $1 (Calculated)"
            ws_calcs[f'B{current_row}'].value = f'=IFERROR(B{current_row-1}/B{current_row-2},0)'
            ws_calcs[f'B{current_row}'].number_format = currency_format
            ws_calcs[f'C{current_row}'] = "Pre-Money Valuation / Seed Raise"
            ws_calcs[f'D{current_row}'] = "Calculated from Inputs"
            for col_letter in ['A', 'B', 'C', 'D']:
                 ws_calcs[f'{col_letter}{current_row}'].border = thin_border
            current_row += 2

        elif section["title"] == "Use of Funds Summary":
            for i in range(24, 29):
                ws_calcs[f'A{current_row}'].value = f'=Inputs!A{i}'
                ws_calcs[f'B{current_row}'].value = f'=Inputs!B{i}'
                ws_calcs[f'B{current_row}'].number_format = currency_format
                ws_calcs[f'C{current_row}'].value = f'=Inputs!C{i}'
                ws_calcs[f'C{current_row}'].number_format = percentage_format
                ws_calcs[f'D{current_row}'].value = f'=Inputs!D{i}'
                for col_letter in ['A', 'B', 'C', 'D']:
                     ws_calcs[f'{col_letter}{current_row}'].border = thin_border
                current_row += 1
            ws_calcs[f'A{current_row}'] = "TOTAL"
            ws_calcs[f'B{current_row}'].value = f'=SUM(B{current_row-5}:B{current_row-1})'
            ws_calcs[f'B{current_row}'].number_format = currency_format
            ws_calcs[f'C{current_row}'].value = f'=SUM(C{current_row-5}:C{current_row-1})'
            ws_calcs[f'C{current_row}'].number_format = percentage_format
            for col_letter in ['A', 'B', 'C']:
                ws_calcs[f'{col_letter}{current_row}'].font = Font(bold=True)
                ws_calcs[f'{col_letter}{current_row}'].border = thin_border
            current_row += 2

        elif section["title"] == "Revenue Projections Summary":
            for i in range(33, 37):
                ws_calcs[f'A{current_row}'].value = f'=Inputs!A{i}'
                ws_calcs[f'B{current_row}'].value = f'=Inputs!B{i}'
                ws_calcs[f'B{current_row}'].number_format = currency_format
                ws_calcs[f'C{current_row}'].value = f'=Inputs!C{i}'
                ws_calcs[f'C{current_row}'].number_format = currency_format
                ws_calcs[f'D{current_row}'].value = f'=IFERROR((C{current_row}/B{current_row})-1,0)'
                ws_calcs[f'D{current_row}'].number_format = percentage_format
                for col_letter in ['A', 'B', 'C', 'D']:
                     ws_calcs[f'{col_letter}{current_row}'].border = thin_border
                current_row += 1
            ws_calcs[f'A{current_row}'] = "TOTAL"
            ws_calcs[f'B{current_row}'].value = f'=SUM(B{current_row-4}:B{current_row-1})'
            ws_calcs[f'B{current_row}'].number_format = currency_format
            ws_calcs[f'C{current_row}'].value = f'=SUM(C{current_row-4}:C{current_row-1})'
            ws_calcs[f'C{current_row}'].number_format = currency_format
            ws_calcs[f'D{current_row}'].value = f'=IFERROR((C{current_row}/B{current_row})-1,0)'
            ws_calcs[f'D{current_row}'].number_format = percentage_format
            for col_letter in ['A', 'B', 'C', 'D']:
                ws_calcs[f'{col_letter}{current_row}'].font = Font(bold=True)
                ws_calcs[f'{col_letter}{current_row}'].border = thin_border
            current_row += 2

        elif section["title"] == "Return Analysis":
            for i in range(40, 44):
                ws_calcs[f'A{current_row}'].value = f'=Inputs!A{i}'
                # Equity Value (Y1) = (Year 1 MC / Pre-Launch Valuation) * Seed Raise
                ws_calcs[f'B{current_row}'].value = f'=IFERROR((Inputs!B{i}/Inputs!$B$18)*Inputs!$B$17,0)'
                ws_calcs[f'B{current_row}'].number_format = currency_format
                # Equity Value (Y3) = (Year 3 MC / Pre-Launch Valuation) * Seed Raise
                ws_calcs[f'C{current_row}'].value = f'=IFERROR((Inputs!C{i}/Inputs!$B$18)*Inputs!$B$17,0)'
                ws_calcs[f'C{current_row}'].number_format = currency_format
                # Return Multiple (Calculated) = Equity Value (Y3) / Seed Raise
                ws_calcs[f'D{current_row}'].value = f'=IFERROR(C{current_row}/Inputs!$B$17,0)'
                ws_calcs[f'D{current_row}'].number_format = '0.00"x"'
                for col_letter in ['A', 'B', 'C', 'D']:
                     ws_calcs[f'{col_letter}{current_row}'].border = thin_border
                current_row += 1
            current_row += 2

    # Auto-fit columns for Calculations sheet
    for col in ws_calcs.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_calcs.column_dimensions[column].width = adjusted_width

    # Freeze panes
    ws_calcs.freeze_panes = 'A3' # Freeze rows above the first data set

    # --- Dashboard Sheet ---
    ws_dashboard = wb.create_sheet("Dashboard")
    ws_dashboard['A1'] = "AEQUITAS PROTOCOL - EXECUTIVE DASHBOARD"
    ws_dashboard['A1'].font = Font(bold=True, size=18, color="1F4E79")
    ws_dashboard.merge_cells('A1:H1')
    ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Key Performance Indicators (KPIs)
    kpi_start_row = 3
    ws_dashboard[f'A{kpi_start_row}'] = "Key Performance Indicators"
    ws_dashboard[f'A{kpi_start_row}'].font = Font(bold=True, size=14, color="1F4E79")
    ws_dashboard.merge_cells(f'A{kpi_start_row}:D{kpi_start_row}')

    kpi_headers = ["Metric", "Value", "Unit", "Trend"]
    for col_idx, header in enumerate(kpi_headers):
        cell = ws_dashboard.cell(row=kpi_start_row + 1, column=col_idx + 1)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Seed Raise KPI
    ws_dashboard[f'A{kpi_start_row+2}'] = "Seed Raise"
    ws_dashboard[f'B{kpi_start_row+2}'].value = '=Inputs!B17'
    ws_dashboard[f'B{kpi_start_row+2}'].number_format = currency_format
    ws_dashboard[f'C{kpi_start_row+2}'] = "USD"
    ws_dashboard[f'D{kpi_start_row+2}'] = "Achieved"
    for col_letter in ['A', 'B', 'C', 'D']:
         ws_dashboard[f'{col_letter}{kpi_start_row+2}'].border = thin_border

    # Pre-Money Valuation KPI
    ws_dashboard[f'A{kpi_start_row+3}'] = "Pre-Money Valuation"
    ws_dashboard[f'B{kpi_start_row+3}'].value = '=Inputs!B18'
    ws_dashboard[f'B{kpi_start_row+3}'].number_format = currency_format
    ws_dashboard[f'C{kpi_start_row+3}'] = "USD"
    ws_dashboard[f'D{kpi_start_row+3}'] = "Confirmed"
    for col_letter in ['A', 'B', 'C', 'D']:
         ws_dashboard[f'{col_letter}{kpi_start_row+3}'].border = thin_border

    # Equity Percentage KPI
    ws_dashboard[f'A{kpi_start_row+4}'] = "Equity Percentage"
    ws_dashboard[f'B{kpi_start_row+4}'].value = '=Calculations!B6'
    ws_dashboard[f'B{kpi_start_row+4}'].number_format = percentage_format
    ws_dashboard[f'C{kpi_start_row+4}'] = "%"
    ws_dashboard[f'D{kpi_start_row+4}'] = "Dilution"
    for col_letter in ['A', 'B', 'C', 'D']:
         ws_dashboard[f'{col_letter}{kpi_start_row+4}'].border = thin_border

    # Operational War Chest KPI
    ws_dashboard[f'A{kpi_start_row+5}'] = "Operational War Chest"
    ws_dashboard[f'B{kpi_start_row+5}'].value = '=Inputs!B11'
    ws_dashboard[f'B{kpi_start_row+5}'].number_format = currency_format
    ws_dashboard[f'C{kpi_start_row+5}'] = "USD"
    ws_dashboard[f'D{kpi_start_row+5}'] = "18-Month Runway"
    for col_letter in ['A', 'B', 'C', 'D']:
         ws_dashboard[f'{col_letter}{kpi_start_row+5}'].border = thin_border

    # Total Addressable Market KPI
    ws_dashboard[f'A{kpi_start_row+6}'] = "Total Addressable Market (TAM)"
    ws_dashboard[f'B{kpi_start_row+6}'].value = '=Inputs!B14'
    ws_dashboard[f'B{kpi_start_row+6}'].number_format = currency_format
    ws_dashboard[f'C{kpi_start_row+6}'] = "USD"
    ws_dashboard[f'D{kpi_start_row+6}'] = "Proven Liability"
    for col_letter in ['A', 'B', 'C', 'D']:
         ws_dashboard[f'{col_letter}{kpi_start_row+6}'].border = thin_border

    # Use of Funds Pie Chart
    pie_chart_row = kpi_start_row + 9
    ws_dashboard[f'A{pie_chart_row}'] = "Use of Funds Distribution"
    ws_dashboard[f'A{pie_chart_row}'].font = Font(bold=True, size=12, color="1F4E79")
    ws_dashboard.merge_cells(f'A{pie_chart_row}:D{pie_chart_row}')

    pie = PieChart()
    labels = Reference(ws_calcs, min_col=1, min_row=12, max_row=16)
    data = Reference(ws_calcs, min_col=2, min_row=11, max_row=16)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Use of Funds"
    pie.dLbls = DataLabelList()
    pie.dLbls.showCatName = False
    pie.dLbls.showPercent = True
    pie.dLbls.showVal = True
    pie.dLbls.showLegendKey = False
    ws_dashboard.add_chart(pie, f'A{pie_chart_row+1}')

    # Revenue Projections Bar Chart
    bar_chart_row = pie_chart_row + 18
    ws_dashboard[f'A{bar_chart_row}'] = "Revenue Projections (Y1 vs Y3)"
    ws_dashboard[f'A{bar_chart_row}'].font = Font(bold=True, size=12, color="1F4E79")
    ws_dashboard.merge_cells(f'A{bar_chart_row}:D{bar_chart_row}')

    bar = BarChart()
    bar.type = "col"
    bar.style = 10
    bar.title = "Revenue Streams"
    bar.y_axis.title = "Amount (USD)"
    bar.x_axis.title = "Revenue Stream"

    data = Reference(ws_calcs, min_col=2, min_row=21, max_row=25)
    cats = Reference(ws_calcs, min_col=1, min_row=22, max_row=25)

    series1 = Series(data, title_from_data=True)
    series1.graphicalProperties.solidFill = "4472C4" # Blue
    bar.append(series1)
    bar.set_categories(cats)

    # Add a second series for Year 3
    data2 = Reference(ws_calcs, min_col=3, min_row=21, max_row=25)
    series2 = Series(data2, title_from_data=True)
    series2.graphicalProperties.solidFill = "ED7D31" # Orange
    bar.append(series2)

    ws_dashboard.add_chart(bar, f'A{bar_chart_row+1}')

    # Scenario Toggle (Delight Feature)
    toggle_row = 3
    ws_dashboard[f'F{toggle_row}'] = "Scenario Analysis"
    ws_dashboard[f'F{toggle_row}'].font = Font(bold=True, size=14, color="1F4E79")
    ws_dashboard.merge_cells(f'F{toggle_row}:H{toggle_row}')

    ws_dashboard[f'F{toggle_row+1}'] = "Select Scenario:"
    ws_dashboard[f'F{toggle_row+1}'].font = Font(bold=True)
    ws_dashboard[f'G{toggle_row+1}'] = "Expected" # Default value for the dropdown
    ws_dashboard[f'G{toggle_row+1}'].border = thin_border
    ws_dashboard[f'G{toggle_row+1}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    # Data Validation for dropdown (simulated, as openpyxl doesn't fully support live dropdowns)
    # In a real Excel, this would be Data -> Data Validation -> List -> Source: Inputs!$A$40:$A$43
    ws_dashboard[f'F{toggle_row+2}'] = "Selected Year 1 MC:"
    ws_dashboard[f'F{toggle_row+2}'].font = Font(bold=True)
    ws_dashboard[f'G{toggle_row+2}'].value = '=IFERROR(VLOOKUP(G4,Inputs!A40:D43,2,FALSE),0)'
    ws_dashboard[f'G{toggle_row+2}'].number_format = currency_format
    ws_dashboard[f'G{toggle_row+2}'].border = thin_border

    ws_dashboard[f'F{toggle_row+3}'] = "Selected Year 3 MC:"
    ws_dashboard[f'F{toggle_row+3}'].font = Font(bold=True)
    ws_dashboard[f'G{toggle_row+3}'].value = '=IFERROR(VLOOKUP(G4,Inputs!A40:D43,3,FALSE),0)'
    ws_dashboard[f'G{toggle_row+3}'].number_format = currency_format
    ws_dashboard[f'G{toggle_row+3}'].border = thin_border

    ws_dashboard[f'F{toggle_row+4}'] = "Calculated Return Multiple:"
    ws_dashboard[f'F{toggle_row+4}'].font = Font(bold=True)
    ws_dashboard[f'G{toggle_row+4}'].value = '=IFERROR(VLOOKUP(G4,Calculations!A32:D35,4,FALSE),0)'
    ws_dashboard[f'G{toggle_row+4}'].number_format = '0.00"x"'
    ws_dashboard[f'G{toggle_row+4}'].border = thin_border

    # Auto-fit columns for Dashboard sheet
    for col in ws_dashboard.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws_dashboard.column_dimensions[column].width = adjusted_width

    # Save the workbook
    file_name = "Aequitas_Protocol_Financial_Model.xlsx"
    wb.save(file_name)
    return file_name


if __name__ == "__main__":
    excel_file = create_excel_workbook()
    print(f"Excel workbook '{excel_file}' created successfully.")

