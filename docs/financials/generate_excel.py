
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_excel_workbook():
    wb = openpyxl.Workbook()

    # --- Inputs Sheet ---
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"

    # Header styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws_inputs['A1'] = "AEQUITAS PROTOCOL - FINANCIAL INPUTS"
    ws_inputs['A1'].font = Font(bold=True, size=16)
    ws_inputs.merge_cells('A1:D1')
    ws_inputs['A1'].alignment = Alignment(horizontal='center')

    # Key Financial Metrics
    ws_inputs['A3'] = "Key Financial Metrics"
    ws_inputs['A3'].font = header_font
    ws_inputs['A4'] = "Metric"
    ws_inputs['B4'] = "Value"
    ws_inputs['C4'] = "Unit"
    ws_inputs['D4'] = "Source/Notes"
    for col in ['A', 'B', 'C', 'D']:
        ws_inputs[col + '4'].font = header_font
        ws_inputs[col + '4'].fill = header_fill
        ws_inputs[col + '4'].border = thin_border

    data_metrics = [
        ["Development Cost", 28000000, "USD", "Tri-Agent Validated Consensus"],
        ["Pre-Launch Valuation", 7000000000, "USD", "Tri-Agent Validated Consensus"],
        ["After-Launch (Y1) Valuation", 200000000000, "USD", "Tri-Agent Validated Consensus"],
        ["Operational War Chest", 22000000, "USD", "Tri-Agent Validated Consensus"]
    ]
    for row_idx, row_data in enumerate(data_metrics, start=5):
        ws_inputs.append(row_data)
        for col_idx, cell_value in enumerate(row_data):
            cell = ws_inputs.cell(row=row_idx, column=col_idx + 1)
            cell.border = thin_border
            if col_idx == 1: # Value column
                cell.number_format = '$#,##0'

    # Investment Details
    ws_inputs['A10'] = "Investment Details"
    ws_inputs['A10'].font = header_font
    ws_inputs['A11'] = "Metric"
    ws_inputs['B11'] = "Value"
    ws_inputs['C11'] = "Unit"
    ws_inputs['D11'] = "Source/Notes"
    for col in ['A', 'B', 'C', 'D']:
        ws_inputs[col + '11'].font = header_font
        ws_inputs[col + '11'].fill = header_fill
        ws_inputs[col + '11'].border = thin_border

    data_investment = [
        ["Seed Raise", 22000000, "USD", ""],
        ["Pre-Money Valuation", 7000000000, "USD", ""],
        ["Equity Percentage", 0.00314, "%", "Calculated: Seed Raise / Pre-Money Valuation"]
    ]
    for row_idx, row_data in enumerate(data_investment, start=12):
        ws_inputs.append(row_data)
        for col_idx, cell_value in enumerate(row_data):
            cell = ws_inputs.cell(row=row_idx, column=col_idx + 1)
            cell.border = thin_border
            if col_idx == 1: # Value column
                if row_data[0] == "Equity Percentage":
                    cell.number_format = '0.000%'
                else:
                    cell.number_format = '$#,##0'
    
    # Use of Funds
    ws_inputs['A16'] = "Use of Funds - 18 Month Runway"
    ws_inputs['A16'].font = header_font
    ws_inputs['A17'] = "Category"
    ws_inputs['B17'] = "Amount"
    ws_inputs['C17'] = "% of Raise"
    ws_inputs['D17'] = "Justification"
    for col in ['A', 'B', 'C', 'D']:
        ws_inputs[col + '17'].font = header_font
        ws_inputs[col + '17'].fill = header_fill
        ws_inputs[col + '17'].border = thin_border

    data_funds = [
        ["Legal & Enforcement", 7500000, 0.341, "'Arbitral Swarm' against initial defendants"],
        ["Security Operations", 5000000, 0.227, "State-level physical/digital protection"],
        ["Elite Core Team", 3000000, 0.136, "AI engineers, legal strategists, ops"],
        ["AI Infrastructure", 2000000, 0.091, "Cerberus Auditor compute, training, data"],
        ["Contingency Reserve", 4500000, 0.205, "Counter-attack defense, legal surprises"]
    ]
    for row_idx, row_data in enumerate(data_funds, start=18):
        ws_inputs.append(row_data)
        for col_idx, cell_value in enumerate(row_data):
            cell = ws_inputs.cell(row=row_idx, column=col_idx + 1)
            cell.border = thin_border
            if col_idx == 1: # Amount column
                cell.number_format = '$#,##0'
            elif col_idx == 2: # % of Raise column
                cell.number_format = '0.0%'
    
    # Total row for Use of Funds
    ws_inputs['A23'] = "TOTAL"
    ws_inputs['B23'].value = "=SUM(B18:B22)"
    ws_inputs['B23'].number_format = '$#,##0'
    ws_inputs['C23'].value = "=SUM(C18:C22)"
    ws_inputs['C23'].number_format = '0.0%'
    for col in ['A', 'B', 'C']:
        ws_inputs[col + '23'].font = Font(bold=True)
        ws_inputs[col + '23'].border = thin_border

    # Return Projections
    ws_inputs['A25'] = "Return Projections - Conservative to Aggressive"
    ws_inputs['A25'].font = header_font
    ws_inputs['A26'] = "Scenario"
    ws_inputs['B26'] = "Year 1 MC"
    ws_inputs['C26'] = "Year 3 MC"
    ws_inputs['D26'] = "Investor Return Multiple"
    for col in ['A', 'B', 'C', 'D']:
        ws_inputs[col + '26'].font = header_font
        ws_inputs[col + '26'].fill = header_fill
        ws_inputs[col + '26'].border = thin_border

    data_returns = [
        ["Conservative", 100000000000, 500000000000, 14],
        ["Expected", 200000000000, 1000000000000, 29],
        ["Aggressive", 300000000000, 3000000000000, 43],
        ["Paradigm Shift", 500000000000, 5000000000000, 71]
    ]
    for row_idx, row_data in enumerate(data_returns, start=27):
        ws_inputs.append(row_data)
        for col_idx, cell_value in enumerate(row_data):
            cell = ws_inputs.cell(row=row_idx, column=col_idx + 1)
            cell.border = thin_border
            if col_idx == 1 or col_idx == 2: # MC columns
                cell.number_format = '$#,##0'
            elif col_idx == 3: # Multiple column
                cell.number_format = '0x'

    # Comparable Analysis
    ws_inputs['A32'] = "Comparable Analysis"
    ws_inputs['A32'].font = header_font
    ws_inputs['A33'] = "Protocol"
    ws_inputs['B33'] = "Function"
    ws_inputs['C33'] = "Market Cap"
    ws_inputs['D33'] = "TAM"
    ws_inputs['E33'] = "Aequitas Advantage"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_inputs[col + '33'].font = header_font
        ws_inputs[col + '33'].fill = header_fill
        ws_inputs[col + '33'].border = thin_border

    data_comparables = [
        ["Chainlink", "Oracle", 8000000000, 1000000000000, "Justice Oracle"],
        ["Filecoin", "Storage", 3000000000, 500000000000, "Legal Data Lake"],
        ["Aequitas", "Enforcement", 7000000000, 131000000000000, "Sovereign AI Enforcement Layer"]
    ]
    for row_idx, row_data in enumerate(data_comparables, start=34):
        ws_inputs.append(row_data)
        for col_idx, cell_value in enumerate(row_data):
            cell = ws_inputs.cell(row=row_idx, column=col_idx + 1)
            cell.border = thin_border
            if col_idx == 2 or col_idx == 3: # Market Cap / TAM columns
                cell.number_format = '$#,##0'

    # Auto-fit columns
    for col in ws_inputs.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_inputs.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Freeze panes for better navigation
    ws_inputs.freeze_panes = 'A5'

    # --- Calculations Sheet ---
    ws_calcs = wb.create_sheet("Calculations")
    ws_calcs['A1'] = "AEQUITAS PROTOCOL - FINANCIAL CALCULATIONS"
    ws_calcs['A1'].font = Font(bold=True, size=16)
    ws_calcs.merge_cells('A1:D1')
    ws_calcs['A1'].alignment = Alignment(horizontal='center')

    # Investment Math
    ws_calcs['A3'] = "Investment Math"
    ws_calcs['A3'].font = header_font
    ws_calcs['A4'] = "Metric"
    ws_calcs['B4'] = "Value"
    ws_calcs['C4'] = "Formula"
    for col in ['A', 'B', 'C']:
        ws_calcs[col + '4'].font = header_font
        ws_calcs[col + '4'].fill = header_fill
        ws_calcs[col + '4'].border = thin_border

    ws_calcs['A5'] = "Seed Raise"
    ws_calcs['B5'].value = "=Inputs!B12"
    ws_calcs['B5'].number_format = '$#,##0'
    ws_calcs['C5'].value = "=Inputs!B12"
    ws_calcs['C5'].font = Font(italic=True)
    ws_calcs['A6'] = "Pre-Money Valuation"
    ws_calcs['B6'].value = "=Inputs!B13"
    ws_calcs['B6'].number_format = '$#,##0'
    ws_calcs['C6'].value = "=Inputs!B13"
    ws_calcs['C6'].font = Font(italic=True)
    ws_calcs['A7'] = "Equity Percentage Offered"
    ws_calcs['B7'].value = "=IFERROR(B5/B6,0)"
    ws_calcs['B7'].number_format = '0.000%'
    ws_calcs['C7'].value = "=B5/B6"
    ws_calcs['C7'].font = Font(italic=True)
    ws_calcs['A8'] = "Post-Money Valuation"
    ws_calcs['B8'].value = "=B6+B5"
    ws_calcs['B8'].number_format = '$#,##0'
    ws_calcs['C8'].value = "=B6+B5"
    ws_calcs['C8'].font = Font(italic=True)
    for row_idx in range(5,9):
        for col_idx in range(1,4):
            ws_calcs.cell(row=row_idx, column=col_idx).border = thin_border

    # Use of Funds Breakdown Check
    ws_calcs['A10'] = "Use of Funds Breakdown Check"
    ws_calcs['A10'].font = header_font
    ws_calcs['A11'] = "Category"
    ws_calcs['B11'] = "Amount"
    ws_calcs['C11'] = "% of Raise (Calculated)"
    ws_calcs['D11'] = "% of Raise (Input)"
    ws_calcs['E11'] = "Variance"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_calcs[col + '11'].font = header_font
        ws_calcs[col + '11'].fill = header_fill
        ws_calcs[col + '11'].border = thin_border

    for row_idx in range(18, 23):
        ws_calcs.cell(row=row_idx-5, column=1).value = f"=Inputs!A{row_idx}"
        ws_calcs.cell(row=row_idx-5, column=2).value = f"=Inputs!B{row_idx}"
        ws_calcs.cell(row=row_idx-5, column=2).number_format = '$#,##0'
        ws_calcs.cell(row=row_idx-5, column=3).value = f"=IFERROR(B{row_idx-5}/$B$5,0)"
        ws_calcs.cell(row=row_idx-5, column=3).number_format = '0.0%'
        ws_calcs.cell(row=row_idx-5, column=4).value = f"=Inputs!C{row_idx}"
        ws_calcs.cell(row=row_idx-5, column=4).number_format = '0.0%'
        ws_calcs.cell(row=row_idx-5, column=5).value = f"=C{row_idx-5}-D{row_idx-5}"
        ws_calcs.cell(row=row_idx-5, column=5).number_format = '0.0%'
        for col_idx in range(1,6):
            ws_calcs.cell(row=row_idx-5, column=col_idx).border = thin_border
    
    ws_calcs['A18'] = "TOTAL"
    ws_calcs['B18'].value = "=SUM(B12:B16)"
    ws_calcs['B18'].number_format = '$#,##0'
    ws_calcs['C18'].value = "=SUM(C12:C16)"
    ws_calcs['C18'].number_format = '0.0%'
    ws_calcs['D18'].value = "=SUM(D12:D16)"
    ws_calcs['D18'].number_format = '0.0%'
    ws_calcs['E18'].value = "=SUM(E12:E16)"
    ws_calcs['E18'].number_format = '0.0%'
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws_calcs[col + '18'].font = Font(bold=True)
        ws_calcs[col + '18'].border = thin_border

    # Investor Return Calculation
    ws_calcs['A20'] = "Investor Return Calculation"
    ws_calcs['A20'].font = header_font
    ws_calcs['A21'] = "Scenario"
    ws_calcs['B21'] = "Year 3 MC"
    ws_calcs['C21'] = "Equity Percentage"
    ws_calcs['D21'] = "Investor Share (Year 3)"
    ws_calcs['E21'] = "Initial Investment"
    ws_calcs['F21'] = "Return Multiple (Calculated)"
    ws_calcs['G21'] = "Return Multiple (Input)"
    ws_calcs['H21'] = "Variance"
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_calcs[col + '21'].font = header_font
        ws_calcs[col + '21'].fill = header_fill
        ws_calcs[col + '21'].border = thin_border

    for row_idx in range(27, 31):
        ws_calcs.cell(row=row_idx-5, column=1).value = f"=Inputs!A{row_idx}"
        ws_calcs.cell(row=row_idx-5, column=2).value = f"=Inputs!C{row_idx}"
        ws_calcs.cell(row=row_idx-5, column=2).number_format = '$#,##0'
        ws_calcs.cell(row=row_idx-5, column=3).value = "=Calculations!B7"
        ws_calcs.cell(row=row_idx-5, column=3).number_format = '0.000%'
        ws_calcs.cell(row=row_idx-5, column=4).value = f"=IFERROR(B{row_idx-5}*C{row_idx-5},0)"
        ws_calcs.cell(row=row_idx-5, column=4).number_format = '$#,##0'
        ws_calcs.cell(row=row_idx-5, column=5).value = "=Calculations!B5"
        ws_calcs.cell(row=row_idx-5, column=5).number_format = '$#,##0'
        ws_calcs.cell(row=row_idx-5, column=6).value = f"=IFERROR(D{row_idx-5}/E{row_idx-5},0)"
        ws_calcs.cell(row=row_idx-5, column=6).number_format = '0.0x'
        ws_calcs.cell(row=row_idx-5, column=7).value = f"=Inputs!D{row_idx}"
        ws_calcs.cell(row=row_idx-5, column=7).number_format = '0x'
        ws_calcs.cell(row=row_idx-5, column=8).value = f"=F{row_idx-5}-G{row_idx-5}"
        ws_calcs.cell(row=row_idx-5, column=8).number_format = '0.0x'
        for col_idx in range(1,9):
            ws_calcs.cell(row=row_idx-5, column=col_idx).border = thin_border

    # Auto-fit columns
    for col in ws_calcs.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_calcs.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Freeze panes
    ws_calcs.freeze_panes = 'A5'

    # --- Outputs Sheet ---
    ws_outputs = wb.create_sheet("Outputs")
    ws_outputs['A1'] = "AEQUITAS PROTOCOL - EXECUTIVE SUMMARY"
    ws_outputs['A1'].font = Font(bold=True, size=16)
    ws_outputs.merge_cells('A1:D1')
    ws_outputs['A1'].alignment = Alignment(horizontal='center')

    # Key Financial Highlights
    ws_outputs['A3'] = "Key Financial Highlights"
    ws_outputs['A3'].font = header_font
    ws_outputs['A4'] = "Metric"
    ws_outputs['B4'] = "Value"
    ws_outputs['C4'] = "Notes"
    for col in ['A', 'B', 'C']:
        ws_outputs[col + '4'].font = header_font
        ws_outputs[col + '4'].fill = header_fill
        ws_outputs[col + '4'].border = thin_border

    ws_outputs['A5'] = "Development Cost"
    ws_outputs['B5'].value = "=Inputs!B5"
    ws_outputs['B5'].number_format = '$#,##0'
    ws_outputs['C5'] = "Technology built at $0 cost (as per pitch)"
    ws_outputs['A6'] = "Pre-Launch Valuation"
    ws_outputs['B6'].value = "=Inputs!B6"
    ws_outputs['B6'].number_format = '$#,##0'
    ws_outputs['C6'] = "Defensible valuation based on triangulated consensus"
    ws_outputs['A7'] = "Operational War Chest"
    ws_outputs['B7'].value = "=Inputs!B8"
    ws_outputs['B7'].number_format = '$#,##0'
    ws_outputs['C7'] = "Required for 18 months of global enforcement operations"
    ws_outputs['A8'] = "Equity Offered"
    ws_outputs['B8'].value = "=Calculations!B7"
    ws_outputs['B8'].number_format = '0.000%'
    ws_outputs['C8'] = "For a $22M seed raise"
    ws_outputs['A9'] = "Expected Return Multiple (Year 3)"
    ws_outputs['B9'].value = "=VLOOKUP(\"Expected\",Calculations!A22:H25,6,FALSE)"
    ws_outputs['B9'].number_format = '0x'
    ws_outputs['C9'] = "Based on expected scenario"
    for row_idx in range(5,10):
        for col_idx in range(1,4):
            ws_outputs.cell(row=row_idx, column=col_idx).border = thin_border

    # Use of Funds Summary
    ws_outputs['A11'] = "Use of Funds Summary"
    ws_outputs['A11'].font = header_font
    ws_outputs['A12'] = "Category"
    ws_outputs['B12'] = "Amount"
    ws_outputs['C12'] = "% of Raise"
    for col in ['A', 'B', 'C']:
        ws_outputs[col + '12'].font = header_font
        ws_outputs[col + '12'].fill = header_fill
        ws_outputs[col + '12'].border = thin_border
    
    for row_idx in range(18,23):
        ws_outputs.cell(row=row_idx-5, column=1).value = f"=Inputs!A{row_idx}"
        ws_outputs.cell(row=row_idx-5, column=2).value = f"=Inputs!B{row_idx}"
        ws_outputs.cell(row=row_idx-5, column=2).number_format = '$#,##0'
        ws_outputs.cell(row=row_idx-5, column=3).value = f"=Inputs!C{row_idx}"
        ws_outputs.cell(row=row_idx-5, column=3).number_format = '0.0%'
        for col_idx in range(1,4):
            ws_outputs.cell(row=row_idx-5, column=col_idx).border = thin_border
    ws_outputs['A18'] = "TOTAL"
    ws_outputs['B18'].value = "=Inputs!B23"
    ws_outputs['B18'].number_format = '$#,##0'
    ws_outputs['C18'].value = "=Inputs!C23"
    ws_outputs['C18'].number_format = '0.0%'
    for col in ['A', 'B', 'C']:
        ws_outputs[col + '18'].font = Font(bold=True)
        ws_outputs[col + '18'].border = thin_border

    # Light-Delight Feature: Scenario Toggle for Investor Return
    ws_outputs['A20'] = "Investor Return Scenario Analysis"
    ws_outputs['A20'].font = Font(bold=True, size=14)
    ws_outputs.merge_cells('A20:C20')
    ws_outputs['A20'].alignment = Alignment(horizontal='center')

    ws_outputs['A22'] = "Select Scenario:"
    ws_outputs['B22'] = "Expected" # Default value
    ws_outputs['B22'].font = Font(bold=True, color='FF0000') # Highlight for user input
    ws_outputs['B22'].border = thin_border
    ws_outputs['A22'].font = Font(bold=True)
    ws_outputs['C22'] = "(Conservative, Expected, Aggressive, Paradigm Shift)"
    ws_outputs['C22'].font = Font(italic=True, size=10)

    ws_outputs['A24'] = "Selected Year 3 Market Cap:"
    ws_outputs['B24'].value = "=IFERROR(VLOOKUP(B22,Inputs!A27:C30,3,FALSE), \"N/A\")"
    ws_outputs['B24'].number_format = '$#,##0'
    ws_outputs['B24'].font = Font(bold=True)
    ws_outputs['A24'].font = Font(bold=True)
    ws_outputs['B24'].border = thin_border

    ws_outputs['A25'] = "Calculated Investor Return Multiple:"
    ws_outputs['B25'].value = "=IFERROR(VLOOKUP(B22,Calculations!A22:H25,6,FALSE), \"N/A\")"
    ws_outputs['B25'].number_format = '0.0x'
    ws_outputs['B25'].font = Font(bold=True)
    ws_outputs['A25'].font = Font(bold=True)
    ws_outputs['B25'].border = thin_border

    # Auto-fit columns
    for col in ws_outputs.columns:
        max_length = 0
        column = col[0].column # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws_outputs.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Freeze panes
    ws_outputs.freeze_panes = 'A5'

    wb.save("Aequitas_Protocol_Financial_Architecture.xlsx")

if __name__ == '__main__':
    create_excel_workbook()

